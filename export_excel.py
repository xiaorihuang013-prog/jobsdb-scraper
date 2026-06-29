#!/usr/bin/env python3
"""
从 CSV 生成格式化的 Excel 文件（仅前10列）。

用法:
    python3 export_excel.py \
      --input "../工作 - Sheet1.csv" \
      --output "岗位收集_Marketing_20260629.xlsx"

    # 仅导出新增部分（指定起始序号）
    python3 export_excel.py \
      --input "../工作 - Sheet1.csv" \
      --output "新岗位_20260629.xlsx" \
      --from-seq 52
"""

import argparse
import csv
import io
import json
import re
import sys
import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] 需要安装 openpyxl: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# 前10列的表头
HEADERS_10 = [
    "序号", "类型", "行业", "公司", "公司评分",
    "职位名称", "薪资范围", "工作描述", "地点", "职位名称点开的链接",
]

# 列宽预设
COL_WIDTHS = {
    "序号": 6,
    "类型": 10,
    "行业": 12,
    "公司": 28,
    "公司评分": 8,
    "职位名称": 30,
    "薪资范围": 18,
    "工作描述": 60,
    "地点": 18,
    "职位名称点开的链接": 40,
}


def read_csv(filepath: str) -> list:
    """读取 CSV，返回行列表（每条为 dict）。"""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        content = f.read()

    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for row in reader:
        # 跳过完全空行
        if not any(v.strip() for v in row.values() if v):
            continue
        rows.append(row)
    return rows


def generate_excel(rows: list, output_path: str, from_seq: int = 0,
                   title: str = "", sheet_name: str = "岗位收集"):
    """生成格式化的 Excel 文件。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── 样式定义 ──────────────────────────────────────────────────
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    alt_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")

    # ── 标题行（可选） ─────────────────────────────────────────────
    current_row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Microsoft YaHei", size=14, bold=True, color="2F5496")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row = 2

    # ── 表头 ──────────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS_10, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    current_row += 1

    # ── 数据行 ─────────────────────────────────────────────────────
    row_count = 0
    for data_row in rows:
        # 提取序号
        try:
            seq = int(data_row.get("序号", "0").strip())
        except (ValueError, AttributeError):
            seq = 0

        # 过滤：如果指定了起始序号，跳过之前的行
        if from_seq > 0 and seq < from_seq:
            continue

        for col_idx, header in enumerate(HEADERS_10, 1):
            # 映射 CSV 列名到实际字段名
            field_map = {
                "序号": "序号",
                "类型": "类型",
                "行业": "行业",
                "公司": "公司",
                "公司评分": "公司评分",
                "职位名称": "职位名称",
                "薪资范围": "薪资范围",
                "工作描述": "工作描述",
                "地点": "地点",
                "职位名称点开的链接": "职位名称点开的链接",
            }
            field = field_map.get(header, header)
            value = data_row.get(field, "")

            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border

            # 居中的列
            if header in ("序号", "类型", "公司评分", "薪资范围"):
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment

            # 交替行颜色
            if row_count % 2 == 1:
                cell.fill = alt_fill

        current_row += 1
        row_count += 1

    # ── 设置列宽 ──────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS_10, 1):
        width = COL_WIDTHS.get(header, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行（表头）
    freeze_row = 2 if title else 1
    ws.freeze_panes = f"A{freeze_row + 1}"

    # ── 设置行高 ──────────────────────────────────────────────────
    ws.row_dimensions[freeze_row].height = 22  # 表头行高

    # ── 保存 ──────────────────────────────────────────────────────
    wb.save(output_path)
    return row_count


def main():
    parser = argparse.ArgumentParser(description="生成格式化 Excel")
    parser.add_argument("--input", "-i", required=True, help="输入 CSV 文件")
    parser.add_argument("--output", "-o", required=True, help="输出 Excel 文件 (.xlsx)")
    parser.add_argument("--from-seq", type=int, default=0, help="仅导出此序号及之后的行")
    parser.add_argument("--title", help="Excel 标题")
    parser.add_argument("--sheet-name", default="岗位收集", help="工作表名称")
    parser.add_argument("--new-only", help="仅导出新岗位 JSON，生成独立 Excel")

    args = parser.parse_args()

    if args.new_only:
        # 从 JSON 直接生成 Excel（不需要 CSV 中间步骤）
        with open(args.new_only, "r", encoding="utf-8") as f:
            jobs = json.load(f)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = args.sheet_name

        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="Microsoft YaHei", size=10)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # 表头
        for col_idx, h in enumerate(HEADERS_10, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for i, job in enumerate(jobs):
            row_num = i + 2
            values = [
                str(i + 1),
                job.get("company_type", ""),
                job.get("industry_cn", "") or job.get("industry", ""),
                job.get("company", ""),
                job.get("rating", "") or "无评分",
                job.get("title", ""),
                job.get("salary", ""),
                job.get("jd", ""),
                job.get("location", ""),
                job.get("link", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border

        for col_idx, header in enumerate(HEADERS_10, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 15)

        ws.freeze_panes = "A2"
        wb.save(args.output)
        print(f"[DONE] {len(jobs)} 个岗位 → {args.output}", file=sys.stderr)
        return

    # 默认：从 CSV 生成 Excel
    rows = read_csv(args.input)
    print(f"[INFO] 从 CSV 读取 {len(rows)} 行", file=sys.stderr)

    output_path = args.output
    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"

    count = generate_excel(
        rows, output_path,
        from_seq=args.from_seq,
        title=args.title or "",
        sheet_name=args.sheet_name,
    )

    print(f"[DONE] {count} 行 → {output_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
